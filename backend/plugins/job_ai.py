from flask import Blueprint, request, jsonify, send_file
import json
import logging
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from core.dify_client import DifyClient
from core.response import success_response, error_response, parse_dify_outputs

logger = logging.getLogger(__name__)

# Dify 配置（每个智能体独立的 API Key 和 App ID）
DIFY_API_KEY = "app-3fb4da1ee9634e890279034d35b46d3b"
DIFY_API_URL = "http://127.0.0.1:8081/v1/workflows/run"
DIFY_UPLOAD_URL = "http://127.0.0.1:8081/v1/files/upload"

dify_client = DifyClient(DIFY_API_KEY, DIFY_API_URL, DIFY_UPLOAD_URL)

job_ai_bp = Blueprint('job_ai', __name__, url_prefix='/api/job_ai')

@job_ai_bp.route("/analyze", methods=["POST"])
def analyze():
    try:
        file = request.files.get("file")
        inputs = {
            "stage": "analyze",
            "job_name": request.form.get("job_name", ""),
            "industry": request.form.get("industry", ""),
            "description": request.form.get("description", ""),
            "responsibilities": request.form.get("responsibilities", "")
        }
        if file:
            inputs["file"] = dify_client.build_file_input(file.stream, file.filename, file.content_type)
        
        result = dify_client.run_workflow(inputs)
        outputs = parse_dify_outputs(result)
        if "analyze_result" in outputs:
            outputs["analyze_result"] = dify_client.clean_think_tags(outputs["analyze_result"])
        
        return jsonify(success_response(data=outputs))
    except Exception as e:
        logger.error(f"job_ai analyze 异常: {e}", exc_info=True)
        return jsonify(error_response(str(e))), 500

@job_ai_bp.route("/generate", methods=["POST"])
def generate():
    try:
        file = request.files.get("file")
        inputs = {
            "stage": "generate",
            "suggest_job": request.form.get("suggest_job", ""),
            "choose_id": request.form.get("choose_id", ""),
            "job_name": request.form.get("job_name", ""),
            "industry": request.form.get("industry", ""),
            "description": request.form.get("description", ""),
            "responsibilities": request.form.get("responsibilities", ""),
            "file_text": request.form.get("file_text", ""),
            "count": request.form.get("count", "")
        }
        if file:
            inputs["file"] = dify_client.build_file_input(file.stream, file.filename, file.content_type)
        
        result = dify_client.run_workflow(inputs)
        outputs = parse_dify_outputs(result)
        if "generate_result" in outputs:
            outputs["generate_result"] = dify_client.clean_think_tags(outputs["generate_result"])
        
        return jsonify(success_response(data=outputs))
    except Exception as e:
        logger.error(f"job_ai generate 异常: {e}", exc_info=True)
        return jsonify(error_response(str(e))), 500

@job_ai_bp.route("/confirm", methods=["POST"])
def confirm():
    try:
        body = request.get_json()
        inputs = {
            "stage": "confirm",
            "jobs_json": json.dumps(body.get("jobs", []), ensure_ascii=False)
        }
        
        result = dify_client.run_workflow(inputs)
        outputs = parse_dify_outputs(result)
        return jsonify(success_response(data=outputs))
    except Exception as e:
        logger.error(f"job_ai confirm 异常: {e}", exc_info=True)
        return jsonify(error_response(str(e))), 500


@job_ai_bp.route("/export", methods=["POST"])
def export_jobs():
    try:
        body = request.get_json()
        jobs = body.get("jobs", [])
        export_format = body.get("format", "xlsx")

        if not jobs:
            return jsonify(error_response("没有可导出的岗位数据")), 400

        if export_format == "csv":
            return _export_csv(jobs)
        return _export_xlsx(jobs)
    except Exception as e:
        logger.error(f"job_ai export 异常: {e}", exc_info=True)
        return jsonify(error_response(str(e))), 500


def _safe_join(items, sep="、"):
    if not items:
        return ""
    result = []
    for item in items:
        if isinstance(item, str):
            result.append(item)
        elif isinstance(item, dict):
            result.append(item.get("name") or item.get("value") or "")
        else:
            result.append(str(item))
    return sep.join([r for r in result if r])


def _format_competency(items):
    if not items:
        return ""
    result = []
    for item in items:
        if isinstance(item, dict):
            name = item.get("abilityName") or item.get("name") or ""
            level = item.get("level") or ""
            desc = item.get("ruleDescription") or item.get("description") or ""
            parts = [p for p in [name + ("（" + level + "）" if level else ""), desc] if p]
            result.append("：".join(parts) if len(parts) > 1 else parts[0] if parts else "")
        elif isinstance(item, str):
            result.append(item)
        else:
            result.append(str(item))
    return "\n".join([r for r in result if r])


def _export_xlsx(jobs):
    wb = Workbook()
    ws = wb.active
    ws.title = "岗位数据"

    # 表头
    headers = [
        "序号", "岗位名称", "岗位简称", "所属行业", "薪资范围", "专业要求",
        "岗位描述", "工作职责", "任职要求", "相关证书",
        "职业发展-横向", "职业发展-纵向",
        "能力域", "能力点", "岗位能力胜任要求"
    ]

    header_fill = PatternFill(start_color="1677FF", end_color="1677FF", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    for idx, job in enumerate(jobs, 1):
        row = idx + 1
        salary = _safe_join(job.get("salaryRange") or [], " - ")
        resp = _safe_join(job.get("responsibilities") or [], "\n")
        req = _safe_join(job.get("requirements") or [], "\n")
        certs = _safe_join(job.get("certificates") or [], "\n")
        majors = _safe_join(job.get("majors") or [], "、")

        career = job.get("careerPath") or {}
        horiz = _safe_join(career.get("horizontal") or [], "、")
        vert = _safe_join(career.get("vertical") or [], "、")

        ability_domains = _safe_join(job.get("abilityDomains") or [], "\n")
        ability_bindings = _safe_join(job.get("abilityBindings") or [], "\n")
        competency = _format_competency(job.get("competencyConfig") or [])

        values = [
            idx,
            job.get("name", ""),
            job.get("shortName", ""),
            job.get("industry", ""),
            salary,
            majors,
            job.get("description", ""),
            resp,
            req,
            certs,
            horiz,
            vert,
            ability_domains,
            ability_bindings,
            competency
        ]

        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.border = thin_border

    # 调整列宽
    column_widths = [6, 18, 12, 12, 14, 20, 40, 35, 35, 25, 25, 25, 25, 35, 35]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else "A" + chr(64 + i - 26)].width = width

    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"岗位数据_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename
    )


def _export_csv(jobs):
    import csv
    output = io.StringIO()
    writer = csv.writer(output)

    headers = [
        "序号", "岗位名称", "岗位简称", "所属行业", "薪资范围", "专业要求",
        "岗位描述", "工作职责", "任职要求", "相关证书",
        "职业发展-横向", "职业发展-纵向",
        "能力域", "能力点", "岗位能力胜任要求"
    ]
    writer.writerow(headers)

    for idx, job in enumerate(jobs, 1):
        salary = _safe_join(job.get("salaryRange") or [], " - ")
        resp = _safe_join(job.get("responsibilities") or [], "；")
        req = _safe_join(job.get("requirements") or [], "；")
        certs = _safe_join(job.get("certificates") or [], "；")
        majors = _safe_join(job.get("majors") or [], "、")

        career = job.get("careerPath") or {}
        horiz = _safe_join(career.get("horizontal") or [], "、")
        vert = _safe_join(career.get("vertical") or [], "、")

        ability_domains = _safe_join(job.get("abilityDomains") or [], "；")
        ability_bindings = _safe_join(job.get("abilityBindings") or [], "；")
        competency = _format_competency(job.get("competencyConfig") or [])

        writer.writerow([
            idx,
            job.get("name", ""),
            job.get("shortName", ""),
            job.get("industry", ""),
            salary,
            majors,
            job.get("description", ""),
            resp,
            req,
            certs,
            horiz,
            vert,
            ability_domains,
            ability_bindings,
            competency
        ])

    data = output.getvalue().encode("utf-8-sig")
    output = io.BytesIO(data)
    output.seek(0)

    filename = f"岗位数据_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    return send_file(
        output,
        mimetype="text/csv",
        as_attachment=True,
        download_name=filename
    )
