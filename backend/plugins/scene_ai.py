from flask import Blueprint, request, jsonify, send_file
import json
import logging
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from core.dify_client import DifyClient
from core.response import success_response, error_response, parse_dify_outputs

logger = logging.getLogger(__name__)

# Dify 配置（每个智能体独立的 API Key 和 App ID）
DIFY_API_KEY = "app-tQlnaQFWRZ3YPUsKG8dXZNH1"
DIFY_API_URL = "http://127.0.0.1:8081/v1/workflows/run"
DIFY_UPLOAD_URL = "http://127.0.0.1:8081/v1/files/upload"

dify_client = DifyClient(DIFY_API_KEY, DIFY_API_URL, DIFY_UPLOAD_URL)

scene_ai_bp = Blueprint('scene_ai', __name__, url_prefix='/api/scene_ai')


@scene_ai_bp.route("/analyze", methods=["POST"])
def analyze():
    try:
        file = request.files.get("file")
        inputs = {
            "stage": "analyze",
            "scene_name": request.form.get("scene_name", ""),
            "background": request.form.get("background", ""),
            "difficulty": request.form.get("difficulty", ""),
            "industry": request.form.get("industry", ""),
            "majors": request.form.get("majors", "")
        }
        if file:
            inputs["file"] = dify_client.build_file_input(file.stream, file.filename, file.content_type)

        result = dify_client.run_workflow(inputs)
        outputs = parse_dify_outputs(result)
        if "analyze_result" in outputs:
            outputs["analyze_result"] = dify_client.clean_think_tags(outputs["analyze_result"])

        return jsonify(success_response(data=outputs))
    except Exception as e:
        logger.error(f"scene_ai analyze 异常: {e}", exc_info=True)
        return jsonify(error_response(str(e))), 500


@scene_ai_bp.route("/generate", methods=["POST"])
def generate():
    try:
        file = request.files.get("file")
        inputs = {
            "stage": "generate",
            "suggest_task": request.form.get("suggest_task", ""),
            "choose_id": request.form.get("choose_id", ""),
            "scene_name": request.form.get("scene_name", ""),
            "background": request.form.get("background", ""),
            "difficulty": request.form.get("difficulty", ""),
            "industry": request.form.get("industry", ""),
            "majors": request.form.get("majors", ""),
            "file_text": request.form.get("file_text", ""),
            "count": request.form.get("count", "")
        }
        if file:
            inputs["file"] = dify_client.build_file_input(file.stream, file.filename, file.content_type)

        result = dify_client.run_workflow(inputs)
        outputs = parse_dify_outputs(result)
        if "generate_result" in outputs:
            outputs["generate_result"] = dify_client.clean_think_tags(outputs["generate_result"])

        return jsonify(success_response(data=outputs))
    except Exception as e:
        logger.error(f"scene_ai generate 异常: {e}", exc_info=True)
        return jsonify(error_response(str(e))), 500


@scene_ai_bp.route("/confirm", methods=["POST"])
def confirm():
    try:
        body = request.get_json()
        inputs = {
            "stage": "confirm",
            "scenes_json": json.dumps(body.get("scene", {}), ensure_ascii=False)
        }

        result = dify_client.run_workflow(inputs)
        outputs = parse_dify_outputs(result)
        return jsonify(success_response(data=outputs))
    except Exception as e:
        logger.error(f"scene_ai confirm 异常: {e}", exc_info=True)
        return jsonify(error_response(str(e))), 500


@scene_ai_bp.route("/export", methods=["POST"])
def export_scene():
    try:
        body = request.get_json()
        scene = body.get("scene", {})
        export_format = body.get("format", "xlsx")

        tasks = scene.get("tasks", [])
        if not tasks:
            return jsonify(error_response("没有可导出的任务数据")), 400

        if export_format == "csv":
            return _export_csv(scene)
        return _export_xlsx(scene)
    except Exception as e:
        logger.error(f"scene_ai export 异常: {e}", exc_info=True)
        return jsonify(error_response(str(e))), 500


def _safe_join(items, sep="、"):
    if not items:
        return ""
    result = []
    for item in items:
        if isinstance(item, str):
            result.append(item)
        elif isinstance(item, dict):
            result.append(item.get("name") or item.get("value") or "")
        else:
            result.append(str(item))
    return sep.join([r for r in result if r])


def _format_ability_bindings(items):
    if not items:
        return ""
    result = []
    for item in items:
        if isinstance(item, dict):
            name = item.get("name") or ""
            domain = item.get("domain") or ""
            parts = [p for p in [name, domain] if p]
            result.append("（".join(parts) + "）" if len(parts) > 1 else parts[0] if parts else "")
        elif isinstance(item, str):
            result.append(item)
        else:
            result.append(str(item))
    return "\n".join([r for r in result if r])


def _format_resources(items):
    if not items:
        return ""
    result = []
    for item in items:
        if isinstance(item, dict):
            name = item.get("name") or ""
            rtype = item.get("type") or ""
            desc = item.get("description") or ""
            parts = [p for p in [name + ("【" + rtype + "】" if rtype else ""), desc] if p]
            result.append("：".join(parts) if len(parts) > 1 else parts[0] if parts else "")
        elif isinstance(item, str):
            result.append(item)
        else:
            result.append(str(item))
    return "\n".join([r for r in result if r])


def _format_evaluation_standard(std):
    if not std or not isinstance(std, dict):
        return "", ""
    stype = std.get("type", "rubric")
    if stype == "rubric":
        rubric = std.get("rubric", {}) or {}
        dimensions = rubric.get("dimensions", []) or []
        lines = []
        for dim in dimensions:
            if not isinstance(dim, dict):
                continue
            lines.append("【" + (dim.get("name") or "") + "】")
            for level in dim.get("levels", []) or []:
                if isinstance(level, dict):
                    lines.append("  " + (level.get("level") or "") + "：" + (level.get("description") or ""))
        return "评价量规", "\n".join(lines)
    else:
        scoring = std.get("scoring", {}) or {}
        items = scoring.get("items", []) or []
        lines = []
        for item in items:
            if isinstance(item, dict):
                name = item.get("name") or ""
                score = item.get("score", "")
                criteria = item.get("criteria") or ""
                lines.append("【" + name + ("（" + str(score) + "分）" if score != "" else "") + "】" + criteria)
        return "评分规则", "\n".join(lines)


def _export_xlsx(scene):
    wb = Workbook()
    ws = wb.active
    ws.title = "场景任务数据"

    headers = [
        "场景名称", "场景背景简介", "场景难度", "所属行业", "适用专业",
        "任务序号", "任务名称", "任务简介", "任务难度", "学时",
        "任务说明", "考查知识点", "考查能力点", "任务资源",
        "测评形式", "评价标准类型", "评价标准详情"
    ]

    header_fill = PatternFill(start_color="1677FF", end_color="1677FF", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    scene_name = scene.get("name", "")
    background = scene.get("background", "")
    difficulty = scene.get("difficulty", "")
    industry = scene.get("industry", "")
    majors = _safe_join(scene.get("majors") or [], "、")

    for idx, task in enumerate(scene.get("tasks") or [], 1):
        row = idx + 1
        knowledge = _safe_join(task.get("knowledgePoints") or [], "、")
        abilities = _format_ability_bindings(task.get("abilityBindings") or [])
        resources = _format_resources(task.get("resources") or [])
        std_type, std_detail = _format_evaluation_standard(task.get("evaluationStandard"))

        values = [
            scene_name,
            background,
            difficulty,
            industry,
            majors,
            idx,
            task.get("name", ""),
            task.get("description", ""),
            task.get("difficulty", ""),
            task.get("hours", 0),
            task.get("instructions", ""),
            knowledge,
            abilities,
            resources,
            task.get("evaluationForm", ""),
            std_type,
            std_detail
        ]

        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.border = thin_border

    column_widths = [18, 35, 10, 12, 20, 8, 18, 30, 10, 8, 35, 30, 25, 25, 15, 12, 40]
    for i, width in enumerate(column_widths, 1):
        if i <= 26:
            col_letter = chr(64 + i)
        else:
            col_letter = "A" + chr(64 + i - 26)
        ws.column_dimensions[col_letter].width = width

    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"场景任务数据_{scene_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename
    )


def _export_csv(scene):
    import csv
    output = io.StringIO()
    writer = csv.writer(output)

    headers = [
        "场景名称", "场景背景简介", "场景难度", "所属行业", "适用专业",
        "任务序号", "任务名称", "任务简介", "任务难度", "学时",
        "任务说明", "考查知识点", "考查能力点", "任务资源",
        "测评形式", "评价标准类型", "评价标准详情"
    ]
    writer.writerow(headers)

    scene_name = scene.get("name", "")
    background = scene.get("background", "")
    difficulty = scene.get("difficulty", "")
    industry = scene.get("industry", "")
    majors = _safe_join(scene.get("majors") or [], "、")

    for idx, task in enumerate(scene.get("tasks") or [], 1):
        knowledge = _safe_join(task.get("knowledgePoints") or [], "；")
        abilities = _format_ability_bindings(task.get("abilityBindings") or []).replace("\n", "；")
        resources = _format_resources(task.get("resources") or []).replace("\n", "；")
        std_type, std_detail = _format_evaluation_standard(task.get("evaluationStandard"))

        writer.writerow([
            scene_name,
            background,
            difficulty,
            industry,
            majors,
            idx,
            task.get("name", ""),
            task.get("description", ""),
            task.get("difficulty", ""),
            task.get("hours", 0),
            task.get("instructions", ""),
            knowledge,
            abilities,
            resources,
            task.get("evaluationForm", ""),
            std_type,
            std_detail.replace("\n", "；")
        ])

    data = output.getvalue().encode("utf-8-sig")
    output = io.BytesIO(data)
    output.seek(0)

    filename = f"场景任务数据_{scene_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    return send_file(
        output,
        mimetype="text/csv",
        as_attachment=True,
        download_name=filename
    )
